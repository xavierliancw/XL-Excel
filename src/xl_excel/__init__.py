"""Xavier's Python package containing tools he uses to speed up his Excel workflows."""

from pathlib import Path as __Path
from typing import Any as __Any

from openpyxl import load_workbook as __load_workbook
from pandas import DataFrame as __DataFrame

from xl_excel import util as __utils

__version__ = "0.3.0"


class EXDataSetTooLargeToExtract(RuntimeError):
    """In Xavier's opinion, the data selected is too large for this operation."""


def extract_column_as_list(
    fpath_wkbk: __Path, sheet: int | str, col_name: str
) -> list[__Any]:
    """Extracts a column from an Excel workbook as a list. If the column is too big,
    this will fail because Xavier decided that you're doing something wrong if you're
    trying to use data sets that big in this context. Use Dask and use generators
    instead."""
    wkbk = __load_workbook(fpath_wkbk)

    wkst = (
        wkbk[wkbk.sheetnames[sheet]]
        if isinstance(sheet, int)
        else wkbk.get_sheet_by_name(sheet)
    )
    dataframe = __DataFrame(wkst.values)
    pandas_series = dataframe[__utils.pyindex_for(col_name=col_name)]
    if len(pandas_series) > 999:  # Xavier arbitrarily decided this is too big
        raise EXDataSetTooLargeToExtract()
    return list(pandas_series)
